"""
Quality Checks utilities:
  - build_default_checks(dims, table_name) → list of check rows
  - generate_qc_excel(dims, table_name)    → bytes (xlsx)
  - parse_qc_excel(file_bytes)             → list of check dicts
  - parse_custom_checks(cell, col_name, dataos_type) → list of check dicts
  - generate_qc_yaml(checks, table_name, dataset_ref) → str
"""

import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from utils.error_logger import log_error, ErrorCategory, log_quality_check_failure

# ── Colours ───────────────────────────────────────────────────────────────────
COL_HEADER_BG  = "1E3A5F"   # dark blue header
COL_LOCKED_BG  = "E8EDF2"   # light blue-grey for locked columns
COL_EDIT_BG    = "FFFDE7"   # light yellow for editable column
COL_TABLE_BG   = "D6E4F0"   # table-level row tint
COL_WHITE      = "FFFFFF"
COL_ALT        = "F3F6FA"   # alternating row

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Default check logic ───────────────────────────────────────────────────────
def _default_checks_for_col(col_name, dataos_type, is_pk):
    """Return list of check-string tokens for default_checks column."""
    checks = ["missing_count=0"]
    if is_pk:
        checks.append("duplicate_count=0")
    if dataos_type == "time":
        checks.append("freshness=7d")
    if dataos_type == "boolean":
        checks.append("valid_values=true,false")
    return " | ".join(checks)


def build_default_checks(dims, table_name):
    """
    dims: list of dicts with keys name, type, primary_key
    Returns list of row-dicts ready for Excel / display.
    """
    try:
        if not dims or not isinstance(dims, list):
            log_error(ErrorCategory.DATA_VALIDATION_ERROR, "Dimensions must be a non-empty list")
            return []
        
        rows = []

        # Table-level checks always first
        rows.append({
            "column_name":    "(table-level)",
            "dataos_type":    "",
            "default_checks": "row_count=0 | schema",
            "custom_checks":  "",
        })

        for d in dims:
            name    = d.get("name", "").strip()
            dtype   = d.get("type", "string")
            is_pk   = d.get("primary_key", False)
            if not name:
                continue
            rows.append({
                "column_name":    name,
                "dataos_type":    dtype,
                "default_checks": _default_checks_for_col(name, dtype, is_pk),
                "custom_checks":  "",
            })
        return rows
    
    except Exception as e:
        log_error(ErrorCategory.DATA_VALIDATION_ERROR, f"Failed to build default checks: {str(e)}", exception=e)
        return []


# ── Excel generator ───────────────────────────────────────────────────────────
def generate_qc_excel(dims, table_name):
    """Generate a formatted, partially-locked Excel workbook. Returns bytes."""
    try:
        rows = build_default_checks(dims, table_name)
        
        if not rows:
            log_quality_check_failure("generate_qc_excel", "No rows to generate", severity="error")
            return b""

        wb = Workbook()
        ws = wb.active
        ws.title = "Quality Checks"

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 22   # column_name
        ws.column_dimensions["B"].width = 14   # dataos_type
        ws.column_dimensions["C"].width = 48   # default_checks
        ws.column_dimensions["D"].width = 60   # custom_checks

        # ── Header row ────────────────────────────────────────────────────────
        headers = ["column_name", "dataos_type", "default_checks", "custom_checks ✏️"]
        header_fill  = PatternFill("solid", fgColor=COL_HEADER_BG)
        header_font  = Font(bold=True, color=COL_WHITE, name="Arial", size=11)
        locked_fill  = PatternFill("solid", fgColor=COL_LOCKED_BG)
        edit_fill    = PatternFill("solid", fgColor=COL_EDIT_BG)
        locked_font  = Font(name="Arial", size=10, color="444444")
        edit_font    = Font(name="Arial", size=10, color="1A1A1A")

        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = BORDER
        ws.row_dimensions[1].height = 28

        # ── Instruction row ───────────────────────────────────────────────────
        instr_fill = PatternFill("solid", fgColor="FFF9C4")
        instr_font = Font(italic=True, color="555555", name="Arial", size=9)
        instructions = [
            "Do not edit",
            "Do not edit",
            "Pre-filled defaults — do not edit",
            "Add custom checks using pipe | separator.\n"
            "Examples:\n"
            "  missing_count=5\n"
            "  duplicate_count=0\n"
            "  freshness=1d\n"
            "  valid_values=CASH,CARD,UPI\n"
            "  min=0 | max=999999\n"
            "  regex=^[A-Z]{3}$\n"
            "  failed_rows=amount != price * qty",
        ]
        for ci, txt in enumerate(instructions, start=1):
            cell = ws.cell(row=2, column=ci, value=txt)
            cell.font      = instr_font
            cell.fill      = instr_fill
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border    = BORDER
        ws.row_dimensions[2].height = 100

        # ── Data rows ─────────────────────────────────────────────────────────
        for ri, row in enumerate(rows, start=3):
            is_table_level = row["column_name"] == "(table-level)"
            row_bg = COL_TABLE_BG if is_table_level else (COL_WHITE if ri % 2 == 1 else COL_ALT)
            row_fill = PatternFill("solid", fgColor=row_bg)

            for ci, key in enumerate(["column_name", "dataos_type", "default_checks", "custom_checks"], start=1):
                cell = ws.cell(row=ri, column=ci, value=row[key])
                cell.border    = BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

                if ci <= 3:
                    # Locked columns — styled grey
                    cell.fill = PatternFill("solid", fgColor=COL_LOCKED_BG if not is_table_level else COL_TABLE_BG)
                    cell.font = Font(name="Arial", size=10, color="333333")
                else:
                    # Editable custom_checks column
                    cell.fill = PatternFill("solid", fgColor=COL_EDIT_BG)
                    cell.font = edit_font
            ws.row_dimensions[ri].height = 20

        # ── Freeze header rows ────────────────────────────────────────────────
        ws.freeze_panes = "A3"

        # ── Legend sheet ─────────────────────────────────────────────────────
        leg = wb.create_sheet("Legend & Syntax")
        leg.column_dimensions["A"].width = 22
        leg.column_dimensions["B"].width = 55

        leg_header_fill = PatternFill("solid", fgColor=COL_HEADER_BG)
        leg_hdr_font    = Font(bold=True, color=COL_WHITE, name="Arial", size=11)

        leg_rows = [
            ("Keyword",         "What it does"),
            ("missing_count=N", "Null/missing values must be ≤ N  (e.g. missing_count=0)"),
            ("missing_percent=N","Missing % must be ≤ N  (e.g. missing_percent=5)"),
            ("duplicate_count=N","Duplicate rows must be ≤ N  (e.g. duplicate_count=0)"),
            ("duplicate_percent=N","Duplicate % must be ≤ N"),
            ("freshness=Nd",    "Data must not be older than N days  (e.g. freshness=7d)"),
            ("valid_values=A,B,C","Column must only contain listed values (comma-separated)"),
            ("min=N",           "Numeric column minimum value  (e.g. min=0)"),
            ("max=N",           "Numeric column maximum value  (e.g. max=999999)"),
            ("regex=PATTERN",   "Column must match this regex pattern"),
            ("failed_rows=SQL", "Custom SQL condition — rows where condition is TRUE fail"),
            ("",                ""),
            ("Separator",       "Use  |  (pipe) to add multiple checks on one column"),
            ("Example",         "missing_count=0 | valid_values=CASH,CARD,UPI | min=0"),
            ("",                ""),
            ("Notes",           "Do NOT edit column_name, dataos_type or default_checks columns"),
            ("",                "Only fill the custom_checks column (yellow)"),
            ("",                "Do not use | inside failed_rows — use OR instead of ||"),
        ]

        for ri, (kw, desc) in enumerate(leg_rows, start=1):
            is_header = ri == 1
            kw_cell   = leg.cell(row=ri, column=1, value=kw)
            desc_cell = leg.cell(row=ri, column=2, value=desc)
            for cell in (kw_cell, desc_cell):
                cell.border    = BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if is_header:
                    cell.fill = leg_header_fill
                    cell.font = leg_hdr_font
                elif kw == "":
                    cell.fill = PatternFill("solid", fgColor=COL_WHITE)
                    cell.font = Font(name="Arial", size=10)
                else:
                    cell.fill = PatternFill("solid", fgColor=COL_ALT if ri % 2 == 0 else COL_WHITE)
                    cell.font = Font(name="Arial", size=10,
                                     bold=(kw in ("Separator", "Example", "Notes")))
            leg.row_dimensions[ri].height = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
    
    except Exception as e:
        log_error(ErrorCategory.FILE_DOWNLOAD_ERROR, f"Failed to generate QC Excel: {str(e)}", exception=e)
        return b""


# ── Custom check parser ───────────────────────────────────────────────────────
def parse_custom_checks(cell_value, col_name, dataos_type):
    """
    Parse a pipe-separated custom_checks cell into a list of check dicts.
    Each dict: {check_type, col_name, dataos_type, ...params}
    """
    if not cell_value or not str(cell_value).strip():
        return []

    checks = []
    tokens = [t.strip() for t in str(cell_value).split("|") if t.strip()]

    for token in tokens:
        if "=" not in token:
            # bare keyword — skip or handle 'schema'
            continue
        key, _, val = token.partition("=")
        key = key.strip().lower()
        val = val.strip()

        if key == "missing_count":
            checks.append({"check_type": "missing_count", "col_name": col_name,
                            "threshold": val, "dataos_type": dataos_type})
        elif key == "missing_percent":
            checks.append({"check_type": "missing_percent", "col_name": col_name,
                            "threshold": val, "dataos_type": dataos_type})
        elif key == "duplicate_count":
            checks.append({"check_type": "duplicate_count", "col_name": col_name,
                            "threshold": val, "dataos_type": dataos_type})
        elif key == "duplicate_percent":
            checks.append({"check_type": "duplicate_percent", "col_name": col_name,
                            "threshold": val, "dataos_type": dataos_type})
        elif key == "freshness":
            checks.append({"check_type": "freshness", "col_name": col_name,
                            "window": val, "dataos_type": dataos_type})
        elif key == "valid_values":
            values = [v.strip() for v in val.split(",") if v.strip()]
            checks.append({"check_type": "valid_values", "col_name": col_name,
                            "values": values, "dataos_type": dataos_type})
        elif key == "min":
            checks.append({"check_type": "min", "col_name": col_name,
                            "threshold": val, "dataos_type": dataos_type})
        elif key == "max":
            checks.append({"check_type": "max", "col_name": col_name,
                            "threshold": val, "dataos_type": dataos_type})
        elif key == "regex":
            checks.append({"check_type": "regex", "col_name": col_name,
                            "pattern": val, "dataos_type": dataos_type})
        elif key == "failed_rows":
            checks.append({"check_type": "failed_rows", "col_name": col_name,
                            "condition": val, "dataos_type": dataos_type})

    return checks


# ── Excel reader ──────────────────────────────────────────────────────────────
def parse_qc_excel(file_bytes):
    """
    Read uploaded Excel. Returns:
      {"default_rows": [...], "custom_checks": [...parsed check dicts...]}
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    default_rows  = []
    custom_checks = []

    # Skip row 1 (header) and row 2 (instructions) — data starts at row 3
    for row in ws.iter_rows(min_row=3, values_only=True):
        col_name, dataos_type, default_chk, custom_chk = (
            row[0], row[1], row[2], row[3] if len(row) > 3 else None
        )
        if not col_name:
            continue

        col_name = str(col_name).strip()
        dataos_type = str(dataos_type).strip() if dataos_type else ""

        default_rows.append({
            "column_name":    col_name,
            "dataos_type":    dataos_type,
            "default_checks": str(default_chk).strip() if default_chk else "",
            "custom_checks":  str(custom_chk).strip() if custom_chk else "",
        })

        if custom_chk and str(custom_chk).strip():
            parsed = parse_custom_checks(str(custom_chk), col_name, dataos_type)
            custom_checks.extend(parsed)

    return {"default_rows": default_rows, "custom_checks": custom_checks}


# ── Default check dicts (for YAML gen) ───────────────────────────────────────
def _default_check_dicts(default_rows):
    """Convert default_rows back into check dicts for YAML generation."""
    checks = []
    for row in default_rows:
        col   = row["column_name"]
        dtype = row["dataos_type"]
        raw   = row.get("default_checks", "")
        if not raw:
            continue
        tokens = [t.strip() for t in raw.split("|") if t.strip()]
        for token in tokens:
            token_l = token.lower()
            if token_l == "schema":
                checks.append({"check_type": "schema", "col_name": col, "dataos_type": dtype})
            elif "=" in token:
                key, _, val = token.partition("=")
                key = key.strip().lower()
                val = val.strip()
                if key == "row_count":
                    checks.append({"check_type": "row_count", "col_name": col, "threshold": val, "dataos_type": dtype})
                elif key == "missing_count":
                    checks.append({"check_type": "missing_count", "col_name": col, "threshold": val, "dataos_type": dtype})
                elif key == "duplicate_count":
                    checks.append({"check_type": "duplicate_count", "col_name": col, "threshold": val, "dataos_type": dtype})
                elif key == "freshness":
                    checks.append({"check_type": "freshness", "col_name": col, "window": val, "dataos_type": dtype})
                elif key == "valid_values":
                    values = [v.strip() for v in val.split(",") if v.strip()]
                    checks.append({"check_type": "valid_values", "col_name": col, "values": values, "dataos_type": dtype})
    return checks


# ── YAML generator ────────────────────────────────────────────────────────────
def _human_name(col, check_type):
    """Generate a readable check name."""
    if col in ("(table-level)", ""):
        return check_type.replace("_", " ").capitalize()
    pretty = check_type.replace("_", " ")
    return f"{col} should not have {pretty}" if "count" in check_type or "percent" in check_type else f"{col} {pretty}"


def _render_check(c):
    """
    Render a single check dict into Soda YAML lines (indented for the checks list).
    Soda format: the check expression IS the YAML key, sub-keys are name/attributes/etc.

    Returns list of strings, each representing one indented line under 'checks:'.
    Indent base = 16 spaces (inside stackSpec.inputs[].checks).
    """
    ct  = c["check_type"]
    col = c.get("col_name", "")
    ind = "                "   # 16 spaces — aligns under checks:

    def check_block(expr, name, category, extra_lines=None):
        """Build a standard Soda check block."""
        out = []
        out.append(f"{ind}- {expr}:")
        out.append(f"{ind}    name: {name}")
        if extra_lines:
            for l in extra_lines:
                out.append(f"{ind}    {l}")
        out.append(f"{ind}    attributes:")
        out.append(f"{ind}      category: {category}")
        return out

    if ct == "row_count":
        threshold = c.get("threshold", "0")
        expr = f"row_count > {threshold}"
        name = f"{col} row count must be greater than {threshold}" if col not in ("(table-level)", "") else f"Row count must be greater than {threshold}"
        return check_block(expr, name, "Accuracy")

    elif ct == "schema":
        # Schema checks have a special nested structure
        out = []
        out.append(f"{ind}- schema:")
        out.append(f"{ind}    name: Schema validation for {col if col not in ('(table-level)', '') else 'table'}")
        out.append(f"{ind}    attributes:")
        out.append(f"{ind}      category: Schema")
        return out

    elif ct in ("missing_count", "missing_percent"):
        threshold = c.get("threshold", "0")
        op = "=" if threshold == "0" else "<="
        expr = f"{ct}({col}) {op} {threshold}"
        name = f"{col} should not have missing values" if threshold == "0" else f"{col} missing {ct.split('_')[1]} should be <= {threshold}"
        return check_block(expr, name, "Completeness")

    elif ct in ("duplicate_count", "duplicate_percent"):
        threshold = c.get("threshold", "0")
        op = "=" if threshold == "0" else "<="
        expr = f"{ct}({col}) {op} {threshold}"
        name = f"{col} must be unique" if threshold == "0" else f"{col} duplicate {ct.split('_')[1]} should be <= {threshold}"
        return check_block(expr, name, "Uniqueness")

    elif ct == "freshness":
        window = c.get("window", "7d")
        # Soda freshness: freshness(col) < Xd
        unit = window[-1] if window[-1].isalpha() else "d"
        num  = window[:-1] if window[-1].isalpha() else window
        expr = f"freshness({col}) < {num}{unit}"
        name = f"{col} must be fresher than {window}"
        return check_block(expr, name, "Freshness")

    elif ct == "valid_values":
        values = c.get("values", [])
        vals_str = ", ".join(values)
        expr = f"invalid_count({col}) = 0"
        name = f"{col} must only contain valid values"
        extra = [f"valid values: [{vals_str}]"]
        return check_block(expr, name, "Validity", extra)

    elif ct == "min":
        threshold = c.get("threshold", "0")
        expr = f"min({col}) >= {threshold}"
        name = f"{col} minimum value must be >= {threshold}"
        return check_block(expr, name, "Validity")

    elif ct == "max":
        threshold = c.get("threshold", "0")
        expr = f"max({col}) <= {threshold}"
        name = f"{col} maximum value must be <= {threshold}"
        return check_block(expr, name, "Validity")

    elif ct == "regex":
        pattern = c.get("pattern", "")
        expr = f"invalid_count({col}) = 0"
        name = f"{col} must match required format"
        extra = [f"valid regex: '{pattern}'"]
        return check_block(expr, name, "Validity", extra)

    elif ct == "failed_rows":
        condition = c.get("condition", "")
        out = []
        out.append(f"{ind}- failed rows:")
        out.append(f"{ind}    name: {col} failed rows check" if col not in ("(table-level)", "") else f"{ind}    name: Failed rows check")
        out.append(f"{ind}    fail condition: {condition}")
        out.append(f"{ind}    attributes:")
        out.append(f"{ind}      category: Accuracy")
        return out

    return []


def generate_qc_yaml(default_rows, custom_checks, table_name, dataset_ref,
                     qc_name, workspace="public", engine="minerva",
                     cluster_name=""):
    """
    Generate a correct DataOS Soda Quality Check workflow YAML.

    Structure:
      name / version / type / tags / description / workspace
      workflow:
        dag:
          - name: <job>
            spec:
              stack: soda+python:1.0
              compute: runnable-default
              resources: ...
              logLevel: INFO
              stackSpec:
                inputs:
                  - dataset: dataos://<depot>:<collection>/<table>
                    options:
                      engine: minerva
                      clusterName: ...
                    profile:
                      columns:
                        - include *
                    checks:
                      - <check_expr>:
                          name: ...
                          attributes:
                            category: ...
    """
    all_checks = _default_check_dicts(default_rows) + custom_checks

    L = []

    # ── Top-level metadata (name first, per DataOS convention) ─────────────────
    L.append(f"name: {qc_name}")
    L.append("version: v1")
    L.append("type: workflow")
    L.append("tags:")
    L.append("  - workflow")
    L.append("  - soda-checks")
    L.append(f"description: Applying quality checks for the {table_name.upper()} table")
    L.append(f"workspace: {workspace}")
    L.append("")

    # ── Workflow ───────────────────────────────────────────────────────────────
    L.append("workflow:")
    L.append("  dag:")
    L.append(f"    - name: {qc_name}")
    L.append("      spec:")
    L.append("        stack: soda+python:1.0")
    L.append("        compute: runnable-default")
    L.append("        resources:")
    L.append("          requests:")
    L.append("            cpu: 1000m")
    L.append("            memory: 250Mi")
    L.append("          limits:")
    L.append("            cpu: 1000m")
    L.append("            memory: 250Mi")
    L.append("        logLevel: INFO")
    L.append("")
    L.append("        stackSpec:")
    L.append("          inputs:")
    L.append(f"            - dataset: {dataset_ref}")

    # Options block
    if engine or cluster_name:
        L.append("              options:")
        if engine:
            L.append(f"                engine: {engine}")
        if cluster_name:
            L.append(f"                clusterName: {cluster_name}")

    # Profile block
    L.append("              profile:")
    L.append("                columns:")
    L.append("                  - include *")
    L.append("")
    L.append("              checks:")

    # ── Emit each check ────────────────────────────────────────────────────────
    for c in all_checks:
        check_lines = _render_check(c)
        for line in check_lines:
            L.append(line)
        L.append("")   # blank line between checks for readability

    return "\n".join(L)